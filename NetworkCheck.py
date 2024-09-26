

import pandas as pd
import numpy as np
import math
import os

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import filedialog


def browse_raw_data():
    filename = filedialog.askopenfilename(title="Select Raw Data File")
    if filename:
        raw_data_entry.delete(0, tk.END)
        raw_data_entry.insert(0, filename)


def browse_coords():
    filename = filedialog.askopenfilename(title="Select Coordinates File")
    if filename:
        coords_entry.delete(0, tk.END)
        coords_entry.insert(0, filename)


def browse_save_location():
    filename = filedialog.askdirectory(title="Select Save Location")
    if filename:
        save_location_entry.delete(0, tk.END)
        save_location_entry.insert(0, filename)


def main():
    raw_data_path = raw_data_entry.get()
    save_location = save_location_entry.get()
    mainDf = loadExcel(raw_data_path)
    filteredData = filterData(mainDf)
    RAWDataDf = createRawDf(filteredData)
    CoordsDf, CalculationsDf, DifferencesDf = getCoords(RAWDataDf)
    createCalculationsExcel(RAWDataDf, CalculationsDf, DifferencesDf, CoordsDf)
    createFinalFormatedExcel(RAWDataDf, CoordsDf, save_location)


def loadExcel(path: str) -> pd.DataFrame:

    mainDf = pd.read_excel(io=path, header=None)
    mainDf.drop([1, 2, 3, 4, 5, 6, 7],
                axis=1,
                inplace=True)  # Drop useless columns

    mainDf.rename(
        columns={
            0: "StationName",
            8: "Horizontal Distance",
            9: "Height Difference",
            10: "Corrected H Dist",
            11: "Corrected H Diff"
        }, inplace=True)

    return mainDf


def filterData(df: pd.DataFrame) -> list:
    filteredData = dict()

    for i in df.loc[df["StationName"] == "Σταθμός"].index:
        currentStation = df.iloc[i + 1]["StationName"]
        tempIndex = i + 4
        tempDataRows = dict()
        while (True):

            try:
                firstTarget = df.iloc[tempIndex]
            except IndexError:
                break

            if (type(firstTarget["StationName"]) != str):
                break

            if (not firstTarget["StationName"] in tempDataRows):
                tempDataRows[firstTarget["StationName"]] = {
                    "HorizontalDistance": firstTarget["Corrected H Dist"],
                    "HeightDiff": firstTarget["Corrected H Diff"]
                }

            tempIndex += 1
        filteredData[currentStation] = tempDataRows

    return filteredData


def createRawDf(filteredData: list) -> pd.DataFrame:
    RAWData = list()

    for StationKey in filteredData.keys():
        for TargetKey in filteredData[StationKey]:
            RAWData.append([
                StationKey,
                TargetKey,
                filteredData[StationKey][TargetKey]['HorizontalDistance'],
                filteredData[StationKey][TargetKey]['HeightDiff']
            ])

    RAWDataDf = pd.DataFrame(RAWData)
    RAWDataDf[0] = RAWDataDf[0].str.upper()
    RAWDataDf[1] = RAWDataDf[1].str.upper()
    RAWDataDf.set_index([0, 1], inplace=True)
    RAWDataDf.rename_axis(["Stations", "Targets"], inplace=True)
    RAWDataDf.rename(
        columns={
            2: "Horizontal Distance",
            3: "Height Difference"
        },
        inplace=True)

    DropIndexList = list()

    for Station, Target in RAWDataDf.index:
        if ((Target, Station) in RAWDataDf.index) and (not (Station, Target) in DropIndexList):
            DropIndexList.append((Target, Station))

    RAWDataDf.drop(index=DropIndexList, inplace=True)
    RAWDataDf = RAWDataDf.sort_index()

    return RAWDataDf


def getCoords(RAWDataDf: pd.DataFrame) -> pd.DataFrame:

    coords_path = coords_entry.get()

    fullPath = coords_path

    coords = pd.read_csv(
        fullPath,
        header=None,
        index_col=0
    )
    coords.rename_axis(["Stations"], inplace=True)
    coords.rename(columns={1: "X", 2: "Y", 3: "Z"}, inplace=True)

    usedCoords = list()
    calcHorizontalDistance = list()
    calcHeightDiff = list()

    for (Station, Target) in RAWDataDf.index:
        if (not Station in usedCoords):
            usedCoords.append(Station)

        if (not Target in usedCoords):
            usedCoords.append(Target)

        pointA = [coords.loc[Station]['X'], coords.loc[Station]['Y']]
        pointB = [coords.loc[Target]['X'], coords.loc[Target]['Y']]

        dist = round(math.dist(pointA, pointB), 4)
        hDiff = round(coords.loc[Target]['Z'] - coords.loc[Station]['Z'], 4)
        calcHorizontalDistance.append(dist)
        calcHeightDiff.append(hDiff)

    DataCoordsDf = coords.loc[usedCoords].sort_index()

    calcValuesDf, differencesDf = calculate(
        RAWDataDf, calcHorizontalDistance, calcHeightDiff)

    return (DataCoordsDf, calcValuesDf, differencesDf)


def calculate(RAWDataDf: pd.DataFrame, distanceList: list, heightDiffList: list) -> pd.DataFrame:
    calcValues = {
        "Horizontal Distance": distanceList,
        "Height Difference": heightDiffList
    }
    calcValuesDf = pd.DataFrame(calcValues, index=RAWDataDf.index)
    calcValuesDf = calcValuesDf.sort_index()

    differencesDf = round((RAWDataDf - calcValuesDf), 3).sort_index()

    return (calcValuesDf, differencesDf)


def createCalculationsExcel(RAWDataDf: pd.DataFrame, calcValuesDf: pd.DataFrame, differencesDf: pd.DataFrame, DataCoordsDf: pd.DataFrame) -> str:

    save_location = save_location_entry.get()

    fullSavePath = save_location + "\\Test A.xlsx"

    with pd.ExcelWriter(fullSavePath) as writer:
        RAWDataDf.to_excel(writer, sheet_name="FilteredRAWData")
        calcValuesDf.to_excel(writer, sheet_name="CalculatedValues")
        differencesDf.to_excel(writer, sheet_name="TotalDifferences")
        DataCoordsDf.to_excel(writer, sheet_name="UsedCoords")

    return fullSavePath


def createFinalFormatedExcel(RAWDataDf: pd.DataFrame, DataCoordsDf: pd.DataFrame, path: str):
    # Formating
    fullSavePath = path + "\\Test A.xlsx"
    wb = load_workbook(fullSavePath)

    finalWb = Workbook()
    finalSheet = finalWb.active
    finalSheet.title = "Network Check"

    rowOffset = 1
    columnOffset = 1

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for col_num, cell_value in enumerate(row):
                if ((cell_value == None) and (col_num == 0)):
                    cell_value = tempCellValue

                if (col_num == 0):
                    tempCellValue = cell_value

                finalSheet.cell(
                    row=rowOffset + 1,
                    column=(col_num + columnOffset),
                    value=cell_value)
            rowOffset += 1

        rowOffset = 1
        columnOffset += 5

    fullSavePath = path + "\\Test B.xlsx"
    finalWb.save(fullSavePath)

    work = load_workbook(fullSavePath)
    sheet = work.active
    rowOffset = 3

    for Station, Target in RAWDataDf.index:
        for i in range(1, len(DataCoordsDf) + 3):
            if (sheet.cell(row=i, column=16).value == Station):
                tempXa = sheet.cell(row=i, column=17).coordinate
                tempYa = sheet.cell(row=i, column=18).coordinate
                tempZa = sheet.cell(row=i, column=19).coordinate

            if (sheet.cell(row=i, column=16).value == Target):
                tempXb = sheet.cell(row=i, column=17).coordinate
                tempYb = sheet.cell(row=i, column=18).coordinate
                tempZb = sheet.cell(row=i, column=19).coordinate

        dist = f"=SQRT(({tempXa}-{tempXb})^2+({tempYa}-{tempYb})^2)"
        diff = f"={tempZb}-{tempZa}"

        sheet.cell(row=rowOffset, column=8, value=dist)
        sheet.cell(row=rowOffset, column=9, value=diff)

        sheet.cell(row=rowOffset, column=13,
                   value=f"=H{rowOffset}-C{rowOffset}")
        sheet.cell(row=rowOffset, column=14,
                   value=f"=I{rowOffset}-D{rowOffset}")

        rowOffset += 1

    # Insert Row to add Column Titles

    columnOffset = 1
    sheet.cell(row=1, column=columnOffset, value="RAW DATA")
    sheet.cell(row=1, column=columnOffset + 5, value="ΣΥΝΤΕΤΑΓΜΕΝΕΣ")
    sheet.cell(row=1, column=columnOffset + 10, value="ΔΙΑΦΟΡΕΣ")
    sheet.cell(row=1, column=columnOffset + 15, value="ΔΙΚΤΥΟ")

    font = Font(
        name='Calibri',
        size=11,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )

    border = Border(
        left=Side(
            border_style="thick",
            color='FF000000'),
        right=Side(
            border_style="thick",
            color='FF000000'),
        top=Side(
            border_style="thick",
            color='FF000000'),
        bottom=Side(
            border_style="thick",
            color='FF000000'),
    )

    alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0
    )

    coordsColumn = list()

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = alignment

            if (cell.column in coordsColumn):
                cell.number_format = "0.000"
            else:
                cell.number_format = "0.0000"

            if (cell.row == 1 and cell.value != None):
                cell.border = border
                cell.font = font

            if (cell.value == "Horizontal Distance") or (cell.value == "Height Difference"):
                columnLetter = get_column_letter(cell.column)
                sheet.column_dimensions[columnLetter].width = 20
            elif (cell.value == "X") or (cell.value == "Y") or (cell.value == "Z"):
                coordsColumn.append(cell.column)
                columnLetter = get_column_letter(cell.column)
                sheet.column_dimensions[columnLetter].width = 12

    sheet.merge_cells(range_string="A1:D1")
    sheet.merge_cells(range_string="F1:I1")
    sheet.merge_cells(range_string="K1:N1")
    sheet.merge_cells(range_string="P1:S1")

    fullSavePath = path + "\\Network Check.xlsx"
    work.save(fullSavePath)


# Create the main window
root = tk.Tk()
root.title("File Explorer Example")

# Configure grid weights to make the window responsive
root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=3)
root.grid_columnconfigure(2, weight=1)
root.grid_rowconfigure(0, weight=1)
root.grid_rowconfigure(1, weight=1)
root.grid_rowconfigure(2, weight=1)
root.grid_rowconfigure(3, weight=1)

# Raw Data File Path
raw_data_label = tk.Label(root, text="Raw Data File Path")
raw_data_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")
raw_data_entry = tk.Entry(root, width=50)
raw_data_entry.grid(row=0, column=1, padx=10, pady=10, sticky="we")
raw_data_button = tk.Button(root, text="Browse", command=browse_raw_data)
raw_data_button.grid(row=0, column=2, padx=10, pady=10)

# Coordinates File Path
coords_label = tk.Label(root, text="Coordinates File Path")
coords_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")
coords_entry = tk.Entry(root, width=50)
coords_entry.grid(row=1, column=1, padx=10, pady=10, sticky="we")
coords_button = tk.Button(root, text="Browse", command=browse_coords)
coords_button.grid(row=1, column=2, padx=10, pady=10)

# Save Location
save_location_label = tk.Label(root, text="Save Location")
save_location_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")
save_location_entry = tk.Entry(root, width=50)
save_location_entry.grid(row=2, column=1, padx=10, pady=10, sticky="we")
save_location_button = tk.Button(
    root, text="Browse", command=browse_save_location)
save_location_button.grid(row=2, column=2, padx=10, pady=10)

# Execution and Cancel buttons
execute_button = tk.Button(root, text="Execute", command=main)
execute_button.grid(row=3, column=1, padx=10, pady=10, sticky="e")

cancel_button = tk.Button(root, text="Cancel", command=root.quit)
cancel_button.grid(row=3, column=2, padx=10, pady=10, sticky="w")

# Run the application
root.mainloop()

# if __name__ == "__main__":
#     main()
