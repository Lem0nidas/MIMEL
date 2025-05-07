
import pandas as pd
import numpy as np

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkcalendar import DateEntry

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from datetime import datetime
from copy import copy



def check_conditions(func):
    def wrapper(*args, **kwargs):
        if (len(listbox.get(0, tk.END)) == 0):
            messagebox.showwarning(title= "Wrong Input", message="Please choose a Date")
            return

        if (combobox.get() == "Select an option"):
            messagebox.showwarning(title= "Wrong Input", message="Please choose an option from DropDown list (Monitoring Area)")
            return
        return func(*args, **kwargs)
    return wrapper


@check_conditions
def main():
    mainDf = loadExcel()
    copy_to_equations_excel(mainDf)
    messagebox.showinfo(title="Missing Targets", message=f"Destroyed/Missing Targets are: \n{destroyed_targets_list}")
    
    save_directory = save_location_entry.get() + "\\LGOUpdatedMonitoringResults.xlsx"
    wb.save(save_directory)


def loadExcel():
    RawFullPath = raw_data_entry.get()

    df = pd.read_excel(RawFullPath, header=None, index_col=0)
    df.columns = ['X', 'Y', 'Z']
    df.index.name = 'Targets'
    df.index = df.index.map(lambda x: x.upper() if isinstance(x, str) else x)

    return df



def copy_to_equations_excel(rawData):
    global wb
    global destroyed_targets_list
    global measurementTime
    global dates
    global LGORepeatTargets

    LGORepeatTargets = ["LG25", "LG26", "LG41", "LG42", "LG43", "LG44", "LG45", "LG46", "LG47", "LG48", "LG49", "PT2", "PT3"]

    excel_file_path = excel_entry.get()
    measurementTime = combobox.get()
    dates = listbox.get(0, tk.END)


    df = rawData
    wb = load_workbook(excel_file_path)
    
    destroyed_targets_list = list()

    if (measurementTime == "09:00"):
        first_measurement(df)
    else:
        repeats(df)

    return

def first_measurement(rawData):
    global ws
    sheets = wb.sheetnames

    for i in sheets:

        wb.active = wb[i]
        ws = wb.active

        for row_idx in range(ws.max_row, 0, -1):
            row = ws[row_idx]
            if not is_row_empty(row):
                last_non_empty_row = row_idx
                break

        try:
            allDays = rawData.loc[i].values
            index = 0

            for eachDay in allDays:
                if isinstance(eachDay, (list, tuple, np.ndarray)) and len(eachDay) == 3:
                    x, y, z = eachDay

                    process_day(x,y,z, last_non_empty_row, dates[index], i)
                    last_non_empty_row += 1
                    index += 1
                else:
                    x, y, z = allDays

                    process_day(x,y,z, last_non_empty_row, dates[index], i)
                    break

        except KeyError as e:
            destroyed_targets_list.append(e.args[0])
        except Exception as e:
            messagebox.showinfo(title="Error", message=f"An unknown error occured. \nMake sure you upload the correct files. Additionally check that the excel files have the correct format.")
    return

def repeats(rawData):
    global ws
    
    
    for i in LGORepeatTargets:

        wb.active = wb[i]
        ws = wb.active

        for row_idx in range(ws.max_row, 0, -1):
            row = ws[row_idx]
            if not is_row_empty(row):
                last_non_empty_row = row_idx
                break

        try:
            allDays = rawData.loc[i].values
            index = 0

            for eachDay in allDays:
                if isinstance(eachDay, (list, tuple, np.ndarray)) and len(eachDay) == 3:
                    x, y, z = eachDay

                    process_day(x,y,z, last_non_empty_row, dates[index], i)
                    last_non_empty_row += 1
                    index += 1
                else:
                    x, y, z = allDays

                    process_day(x,y,z, last_non_empty_row, dates[index], i)
                    break

        except KeyError as e:
            destroyed_targets_list.append(e.args[0])
        except Exception as e:
            messagebox.showinfo(title="Error", message=f"An unknown error occured. \nMake sure you upload the correct files. Additionally check that the excel files have the correct format.")
            break
    return

def process_day(x, y, z, last_non_empty_row, date, target):
    lastRow = last_non_empty_row
    newRow = lastRow + 1
    ws.insert_rows(newRow)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=lastRow, column=col)
        new_cell = ws.cell(row=newRow, column=col)
        same_time_cell = ws.cell(row=lastRow - 2, column=col)

        if cell.data_type == 'f':
            new_formula = None
            match measurementTime:
                case "09:00" if col == 10 and target in LGORepeatTargets:
                    new_formula = cell.value.replace(str(lastRow), str(newRow)).replace(str(lastRow - 3), str(lastRow))
                case "12:00" if col == 10 and target in LGORepeatTargets:
                    new_formula = cell.value.replace(str(lastRow), str(newRow))
                case "15:00" if col == 10 and target in LGORepeatTargets:
                    new_formula = cell.value.replace(str(lastRow), str(newRow))

            if new_formula is None:
                if str(lastRow - 1) in cell.value:
                    new_formula = cell.value.replace(str(lastRow), str(newRow)).replace(str(lastRow - 1), str(lastRow))
                else: 
                    new_formula = cell.value.replace(str(lastRow), str(newRow))

            if new_formula is not None:
                new_cell.value = new_formula

        elif col == 1:
            new_cell.value = int(cell.value) + 1
        elif col == 2 and target not in destroyed_targets_list:
            new_cell.value = datetime.strptime(date, "%d/%m/%y").date()
        elif col == 3:
            new_cell.value = datetime.strptime(measurementTime, "%H:%M").time()
        else:
            new_cell.value = cell.value

        if cell.has_style:
            new_cell.font = copy(same_time_cell.font)
            new_cell.fill = copy(same_time_cell.fill)
            new_cell.border = copy(cell.border)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

            match measurementTime:
                case "09:00" if col == 15 and target in LGORepeatTargets:
                    new_cell.value = "TPS MEAS/1st ROUND"
                case "12:00" if col == 15 and target in LGORepeatTargets:
                    new_cell.value = "TPS MEAS/2nd ROUND"
                case "15:00" if col == 15 and target in LGORepeatTargets:
                    new_cell.value = "TPS MEAS/3nd ROUND"

    for col, new_value in zip([4, 5, 6], [x, y, z]):
        ws.cell(row=newRow, column=col).value = new_value

    return

def is_row_empty(row):
    """Check if a row contains only NaN or empty values."""
    for cell in row:
        if cell.value not in (None, '', np.nan):
            return False
    return True


def browse_raw_data():
    filename = filedialog.askopenfilename(title="Select Raw Data File")
    if filename:
        raw_data_entry.delete(0, tk.END)
        raw_data_entry.insert(0, filename)

def browse_excel():
    filename = filedialog.askopenfilename(title="Select Excel File")
    if filename:
        excel_entry.delete(0, tk.END)
        excel_entry.insert(0, filename)

def browse_save_location():
    filename = filedialog.askdirectory(title="Select Save Location")
    if filename:
        save_location_entry.delete(0, tk.END)
        save_location_entry.insert(0, filename)


def remove_selected(event):
    selected_index = listbox.curselection()
    if selected_index:
        listbox.delete(selected_index)


def add_date_to_listbox(event):
    selected_date = date_entry.get()

    date_object = datetime.strptime(selected_date, "%m/%d/%y")
    formatted_date = date_object.strftime("%d/%m/%y")

    if formatted_date not in listbox.get(0, tk.END):
        listbox.insert(tk.END, formatted_date)



root = tk.Tk()
root.title("Update Excel file with equations.")


root.geometry("800x300")

root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=3)
root.grid_columnconfigure(2, weight=1)
root.grid_rowconfigure(0, weight=1)
root.grid_rowconfigure(1, weight=1)
root.grid_rowconfigure(2, weight=1)
root.grid_rowconfigure(3, weight=1)
root.grid_rowconfigure(4, weight=1)
root.grid_rowconfigure(5, weight=1)


raw_data_label = tk.Label(root, text="Raw Data File Path")
raw_data_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")
raw_data_entry = tk.Entry(root, width=50)
raw_data_entry.grid(row=0, column=1, padx=10, pady=10, sticky="we")
raw_data_button = tk.Button(root, text="Browse", command=browse_raw_data)
raw_data_button.grid(row=0, column=2, padx=10, pady=10)


excel_label = tk.Label(root, text="Excel File Path")
excel_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")
excel_entry = tk.Entry(root, width=50)
excel_entry.grid(row=1, column=1, padx=10, pady=10, sticky="we")
excel_button = tk.Button(root, text="Browse", command=browse_excel)
excel_button.grid(row=1, column=2, padx=10, pady=10)


save_location_label = tk.Label(root, text="Save Location")
save_location_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")
save_location_entry = tk.Entry(root, width=50)
save_location_entry.grid(row=2, column=1, padx=10, pady=10, sticky="we")
save_location_button = tk.Button(
    root, text="Browse", command=browse_save_location)
save_location_button.grid(row=2, column=2, padx=10, pady=10)


date_label = tk.Label(root, text="Select Date")
date_label.grid(row=3, column=0, padx=10, pady=10, sticky="w")
date_entry = DateEntry(root, width=12, background='darkblue',
                       foreground='white', borderwidth=2)
date_entry.grid(row=3, column=1, padx=10, pady=10, sticky="we")


listbox = tk.Listbox(root, height=3, width=10)
listbox.grid(row=3, column=2, padx=10, pady=10, sticky="we")


measuringTime = ["09:00", "12:00", "15:00"]
dropdown_label = tk.Label(root, text="Time of measurement")
dropdown_label.grid(row=4, column=0, padx=10, pady=10, sticky="w")
combobox = ttk.Combobox(root, values=measuringTime, width=20)
combobox.grid(row=4, column=1, padx=10, pady=10, sticky="we")
combobox.set("Select an option")


date_entry.bind("<<DateEntrySelected>>", add_date_to_listbox)


listbox.bind("<Double-Button-1>", remove_selected)


execute_button = tk.Button(root, text="Execute", command=main)
execute_button.grid(row=5, column=1, padx=10, pady=10, sticky="e")


cancel_button = tk.Button(root, text="Cancel", command=root.quit)
cancel_button.grid(row=5, column=2, padx=10, pady=10, sticky="w")


root.mainloop()
