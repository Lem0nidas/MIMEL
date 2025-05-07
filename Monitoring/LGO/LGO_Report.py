from copy import copy
from openpyxl import load_workbook
from tkinter import filedialog, messagebox, ttk
from tkinterdnd2 import TkinterDnD, DND_FILES

import os
import numpy as np
import tkinter as tk


def check_conditions(func):
    def wrapper(*args, **kwargs):
        if (combobox.get() == "Select an option"):
            messagebox.showwarning(title= "Wrong Input", message="Please choose an option from DropDown list (Number of measurements)")
            return
        return func(*args, **kwargs)
    return wrapper


@check_conditions
def main():
    clean_file_path = file_path_box.get('1.0', 'end').strip().replace('{', '').replace('}', '')
    wb = load_workbook(clean_file_path, data_only=True)
    num_of_lines = int(combobox.get())

    LGORepeatTargets = ["LG25", "LG26", "LG41", "LG42", "LG43", "LG44", "LG45", "LG46", "LG47", "LG48", "LG49", "PT2", "PT3"]
    all_targets_dict = dict()

    # Move to first sheet
    sheets = wb.sheetnames

    for i in sheets:
        k = 1
        wb.active = wb[i]
        ws = wb.active

        # Find the last row with actual data
        for row_idx in range(ws.max_row, 0, -1):  # Iterate from last row to the first
            row = ws[row_idx]
            if not is_row_empty(row):
                last_non_empty_row = row_idx
                break

        all_lines_dict = dict()

        if i in LGORepeatTargets:
            for k in range(num_of_lines, 0, -1):
                row = ws[last_non_empty_row - k + 1]  # Get the k-th row from the bottom

                list_of_row_values = [cell.value for cell in row]
                all_lines_dict[k] = list_of_row_values
            
            all_targets_dict[i] = all_lines_dict
        else:
            row = ws[last_non_empty_row]
            list_of_row_values = [cell.value for cell in row]
            all_targets_dict[i] = list_of_row_values


    wb.close()

    #############
    clean_destination_path = destination_path_box.get('1.0', 'end').strip().replace('{', '').replace('}', '')
    wb = load_workbook(clean_destination_path)

    # Move to first sheet
    sheets = wb.sheetnames

    for key, val in all_targets_dict.items():
        k = 1
        wb.active = wb[key]
        ws = wb.active

        # Find the last row with actual data
        for row_idx in range(ws.max_row, 0, -1):  # Iterate from last row to the first
            row = ws[row_idx]
            if not is_row_empty(row):
                last_non_empty_row = row_idx
                break

        if isinstance(val, dict):
            for k in range(num_of_lines, 0, -1):
                source_row_data = all_targets_dict[key][k]  # Get the k-th row from the dictionary
                max_col = len(source_row_data)

                for col in range(1, max_col + 1):
                    cell = ws.cell(row=last_non_empty_row, column=col)
                    new_cell = ws.cell(row=last_non_empty_row + 1, column=col)

                    new_cell.value = source_row_data[col - 1]

                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.fill = copy(cell.fill)
                        new_cell.border = copy(cell.border)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

                last_non_empty_row += 1
        else:
            date = ws[last_non_empty_row][1].value
            if (val[1] == date):
                continue
            else:
                source_row_data = all_targets_dict[key]
                max_col = len(source_row_data)

                for col in range(1, max_col + 1):
                    cell = ws.cell(row=last_non_empty_row, column=col)
                    new_cell = ws.cell(row=last_non_empty_row + 1, column=col)

                    new_cell.value = source_row_data[col - 1]

                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.fill = copy(cell.fill)
                        new_cell.border = copy(cell.border)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)


    save_directory = save_location_entry.get() + "\\LGO-Report.xlsx"
    wb.save(save_directory)

def is_row_empty(row):
    """Check if a row contains only NaN or empty values."""
    for cell in row:
        if cell.value not in (None, '', np.nan):
            return False
    return True

def browse_origin_excel():
    filename = filedialog.askopenfilename(title="Select Raw Data File")
    if filename:
        file_path_box.delete('1.0', 'end')
        file_path_box.insert('1.0', filename)

def browse_destination_excel():
    filename = filedialog.askopenfilename(title="Select Excel File")
    if filename:
        destination_path_box.delete('1.0', 'end')
        destination_path_box.insert('1.0', filename)

def browse_save_location():
    filename = filedialog.askdirectory(title="Select Save Location")
    if filename:
        save_location_entry.delete(0, tk.END)
        save_location_entry.insert(0, filename)

def on_origin_drop(event):
    file_path_box.delete('1.0', 'end')
    file_path_box.insert('1.0', event.data)
    return

def on_destination_drop(event):
    destination_path_box.delete('1.0', 'end')
    destination_path_box.insert('1.0', event.data)
    return

def autofill_save_location(event=None):
    destination_text = destination_path_box.get("1.0", "end").strip().replace('{', '').replace('}', '')
    folder_path = os.path.dirname(destination_text)
    save_location_entry.delete(0, "end")
    save_location_entry.insert(0, folder_path)


def on_close():
    root.destroy()

# Create the main window
root = TkinterDnD.Tk()
root.title("Update Excel file with equations.")

# Set geometry
root.geometry("800x300")

# Configure grid weights to make the window responsive
root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=3)
root.grid_columnconfigure(2, weight=1)
root.grid_rowconfigure(0, weight=1)
root.grid_rowconfigure(1, weight=1)
root.grid_rowconfigure(2, weight=1)
root.grid_rowconfigure(3, weight=1)
root.grid_rowconfigure(4, weight=1)
root.grid_rowconfigure(5, weight=1)


# Origin Excel
origin_label = tk.Label(root, text="Origin Excel")
origin_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
origin_button = tk.Button(root, text="Browse", command=browse_origin_excel)
origin_button.grid(row=0, column=2, padx=5, pady=5)

file_path_box = tk.Text(root, width=40, height=3, bg="lightgray", wrap='word')
file_path_box.insert('1.0', "Drag and drop files here...")
file_path_box.grid(row=0, column=1, padx=5, pady=5)
file_path_box.drop_target_register(DND_FILES)
file_path_box.dnd_bind('<<Drop>>', on_origin_drop)

# Destination Excel
destination_label = tk.Label(root, text="Destination Excel")
destination_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
destination_button = tk.Button(root, text="Browse", command=lambda: [browse_destination_excel(), autofill_save_location()])
destination_button.grid(row=1, column=2, padx=5, pady=5)

destination_path_box = tk.Text(root, width=40, height=3, bg="lightgray", wrap='word')
destination_path_box.insert('1.0', "Drag and drop files here...")
destination_path_box.grid(row=1, column=1, padx=5, pady=5)
destination_path_box.drop_target_register(DND_FILES)
destination_path_box.dnd_bind("<<Drop>>", lambda event: [on_destination_drop(event), autofill_save_location()])

# Save Location
save_location_label = tk.Label(root, text="Save Location")
save_location_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
save_location_entry = tk.Entry(root, width=30, readonlybackground="lightgray")
save_location_entry.grid(row=2, column=1, padx=5, pady=5, sticky="we")
save_location_button = tk.Button(
    root, text="Browse", command=browse_save_location)
save_location_button.grid(row=2, column=2, padx=5, pady=5)

# Create a Combobox with a few options
monitoringAreas = ["1", "2", "3", "4", "5", "6"]
dropdown_label = tk.Label(root, text="Number of dates/measurements")
dropdown_label.grid(row=4, column=0, padx=5, pady=5, sticky="w")
combobox = ttk.Combobox(root, values=monitoringAreas, width=30)
combobox.grid(row=4, column=1, padx=5, pady=5, sticky="we")
combobox.set("Select an option")  # Default text

# Execution and Cancel buttons
execute_button = tk.Button(root, text="Execute", command=main)
execute_button.grid(row=5, column=1, padx=5, pady=5, sticky="e")

cancel_button = tk.Button(root, text="Cancel", command=root.quit)
cancel_button.grid(row=5, column=2, padx=5, pady=5, sticky="w")

root.protocol("WM_DELETE_WINDOW", on_close)

# Run the application
root.mainloop()