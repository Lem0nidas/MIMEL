import pandas as pd
import numpy as np

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkcalendar import Calendar, DateEntry

from openpyxl import load_workbook
from openpyxl.chart import Reference, Series

from datetime import datetime, timedelta
from copy import copy



def check_conditions(func):
    def wrapper(*args, **kwargs):
        if (len(listbox.get(0, tk.END)) == 0):
            messagebox.showwarning(title= "Wrong Input", message="Please choose a Date")
            return

        if (combobox.get() == "Select an option"):
            messagebox.showwarning(title= "Wrong Input", message="Please choose an option from DropDown list (Monitoring Area)")
            return
        return func(*args, **kwargs)
    return wrapper


@check_conditions
def main():
    mainDf = loadExcel()
    copy_to_equations_excel(mainDf)
    messagebox.showinfo(title="Missing Targets", message=f"Destroyed/Missing Targets are: \n{destroyed_targets_list}")
    
    save_directory = save_location_entry.get() + f"\\{combobox.get()}UpdatedMonitoringResults.xlsx"
    wb.save(save_directory)


def loadExcel():
    RawFullPath = raw_data_entry.get()

    df = pd.read_excel(RawFullPath, header=None, index_col=0)
    df.columns = ['X', 'Y', 'Z']
    df.index.name = 'Targets'

    return df

def copy_to_equations_excel(rawData):
    global wb
    global ws
    global destroyed_targets_list

    excel_file_path = excel_entry.get()
    dates = listbox.get(0, tk.END)

    df = rawData
    wb = load_workbook(excel_file_path)
    destroyed_targets_list = list()


    # Move to first sheet
    sheets = wb.sheetnames


    for i in sheets:

        wb.active = wb[i]
        ws = wb.active

        # Find the last row with actual data
        for row_idx in range(ws.max_row, 0, -1):  # Iterate from last row to the first
            row = ws[row_idx]
            if not is_row_empty(row):
                last_non_empty_row = row_idx
                break

        try:
            allDays = df.loc[i].values
            index = 0

            for eachDay in allDays:
                # Check if 'day' is an iterable and has exactly 3 elements
                if isinstance(eachDay, (list, tuple, np.ndarray)) and len(eachDay) == 3:
                    x, y, z = eachDay

                    process_day(x,y,z, last_non_empty_row, dates[index])
                    last_non_empty_row += 1
                    index += 1
                else:
                    x, y, z = allDays

                    process_day(x,y,z, last_non_empty_row, dates[index])
                    break

        except KeyError as e:
            print(f"The given Target is missing from Data file: {e}")
            destroyed_targets_list.append(e.args[0])

            for date in dates:
                add_destroyed_target(last_non_empty_row, date)
                last_non_empty_row += 1
        except Exception as e:
            print(e)

    return

def process_day(x, y, z, last_non_empty_row, date):
    lastRow = last_non_empty_row
    newRow = lastRow + 1
    ws.insert_rows(newRow)


    # Copy the previous row to the new row, adjusting formulas
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=lastRow, column=col)
        new_cell = ws.cell(row=newRow, column=col)

        if cell.data_type == 'f':  # If the cell contains a formula
            # Adjust the formula to reference the new row
            if str(lastRow - 1) in cell.value:
                new_formula = cell.value.replace(str(lastRow), str(newRow)).replace(str(lastRow - 1), str(lastRow))
            else: 
                new_formula = cell.value.replace(str(lastRow), str(newRow))

            new_cell.value = new_formula
        elif col == 1:
            new_cell.value = int(cell.value) + 1
        elif col == 2:
            new_cell.value = datetime.strptime(date, "%d/%m/%y").date()
        else: 
            new_cell.value = cell.value

        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

    # Update columns D, E, F with new values (ONLY FOR 150 & POND2)
    if (combobox.get() == "150") or (combobox.get() == "POND2"):
        new_values = [x, y, z]
        columns_to_update = [4, 5, 6]  # Columns D, E, F
    else:
        # Update columns C, D, E with new values (EXCEPT FROM 150 & POND2)
        new_values = [x, y, z]
        columns_to_update = [3, 4, 5]  # Columns C, D, E

    for col, new_value in zip(columns_to_update, new_values):
        ws.cell(row=newRow, column=col).value = new_value

    return

def add_destroyed_target(last_non_empty_row, date):
    lastRow = last_non_empty_row
    newRow = lastRow + 1
    ws.insert_rows(newRow)


    # Copy the previous row to the new row, adjusting formulas
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=lastRow, column=col)
        new_cell = ws.cell(row=newRow, column=col)

        if cell.data_type == 'f':  # If the cell contains a formula
            # Adjust the formula to reference the new row
            new_formula = cell.value.replace(str(lastRow), str(newRow))
            new_cell.value = new_formula
        elif col == 1:
            new_cell.value = int(cell.value) + 1
        elif col == 2:
            new_cell.value = datetime.strptime(date, "%d/%m/%y").date()
        else: 
            new_cell.value = cell.value

        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)
    
    return


def is_row_empty(row):
    """Check if a row contains only NaN or empty values."""
    for cell in row:
        if cell.value not in (None, '', np.nan):  # Check for non-empty values
            return False
    return True


def browse_raw_data():
    filename = filedialog.askopenfilename(title="Select Raw Data File")
    if filename:
        raw_data_entry.delete(0, tk.END)
        raw_data_entry.insert(0, filename)

def browse_excel():
    filename = filedialog.askopenfilename(title="Select Excel File")
    if filename:
        excel_entry.delete(0, tk.END)
        excel_entry.insert(0, filename)

def browse_save_location():
    filename = filedialog.askdirectory(title="Select Save Location")
    if filename:
        save_location_entry.delete(0, tk.END)
        save_location_entry.insert(0, filename)

# Function to handle double-click event and remove item from Listbox
def remove_selected(event):
    selected_index = listbox.curselection()  # Get the index of the selected item
    if selected_index:  # Check if an item is selected
        listbox.delete(selected_index)  # Remove the selected item

# Function to handle adding selected date to Listbox
def add_date_to_listbox(event):
    selected_date = date_entry.get()  # Get the selected date from DateEntry

    # Convert the date from mm/dd/yy to dd/mm/yy
    date_object = datetime.strptime(selected_date, "%m/%d/%y")
    formatted_date = date_object.strftime("%d/%m/%y")

    if formatted_date not in listbox.get(0, tk.END):  # Check if date is not already in listbox
        listbox.insert(tk.END, formatted_date)  # Insert the date at the end of the listbox


# Create the main window
root = tk.Tk()
root.title("Update Excel file with equations.")

# Set geometry
root.geometry("800x300")

# Configure grid weights to make the window responsive
root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=3)
root.grid_columnconfigure(2, weight=1)
root.grid_rowconfigure(0, weight=1)
root.grid_rowconfigure(1, weight=1)
root.grid_rowconfigure(2, weight=1)
root.grid_rowconfigure(3, weight=1)
root.grid_rowconfigure(4, weight=1)
root.grid_rowconfigure(5, weight=1)


# Raw Data File Path
raw_data_label = tk.Label(root, text="Raw Data File Path")
raw_data_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")
raw_data_entry = tk.Entry(root, width=50)
raw_data_entry.grid(row=0, column=1, padx=10, pady=10, sticky="we")
raw_data_button = tk.Button(root, text="Browse", command=browse_raw_data)
raw_data_button.grid(row=0, column=2, padx=10, pady=10)

# Excel File Path
excel_label = tk.Label(root, text="Excel File Path")
excel_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")
excel_entry = tk.Entry(root, width=50)
excel_entry.grid(row=1, column=1, padx=10, pady=10, sticky="we")
excel_button = tk.Button(root, text="Browse", command=browse_excel)
excel_button.grid(row=1, column=2, padx=10, pady=10)

# Save Location
save_location_label = tk.Label(root, text="Save Location")
save_location_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")
save_location_entry = tk.Entry(root, width=50)
save_location_entry.grid(row=2, column=1, padx=10, pady=10, sticky="we")
save_location_button = tk.Button(
    root, text="Browse", command=browse_save_location)
save_location_button.grid(row=2, column=2, padx=10, pady=10)

# Date Picker
date_label = tk.Label(root, text="Select Date")
date_label.grid(row=3, column=0, padx=10, pady=10, sticky="w")
date_entry = DateEntry(root, width=12, background='darkblue',
                       foreground='white', borderwidth=2)
date_entry.grid(row=3, column=1, padx=10, pady=10, sticky="we")

# Listbox next to DateEntry in the third row
listbox = tk.Listbox(root, height=3, width=10)  # Small Listbox
listbox.grid(row=3, column=2, padx=10, pady=10, sticky="we")

# Create a Combobox with a few options
monitoringAreas = ["129", "131", "134", "136", "137", "138", "150", "KT2", "LGO", "POND2"]
dropdown_label = tk.Label(root, text="Monitoring Area")
dropdown_label.grid(row=4, column=0, padx=10, pady=10, sticky="w")
combobox = ttk.Combobox(root, values=monitoringAreas, width=20)
combobox.grid(row=4, column=1, padx=10, pady=10, sticky="we")
combobox.set("Select an option")  # Default text

# Bind the DateEntry selection event to add the selected date to the Listbox
date_entry.bind("<<DateEntrySelected>>", add_date_to_listbox)

# Bind the double-click event to the Listbox
listbox.bind("<Double-Button-1>", remove_selected)

# Execution and Cancel buttons
execute_button = tk.Button(root, text="Execute", command=main)
execute_button.grid(row=5, column=1, padx=10, pady=10, sticky="e")


cancel_button = tk.Button(root, text="Cancel", command=root.quit)
cancel_button.grid(row=5, column=2, padx=10, pady=10, sticky="w")


# Run the application
root.mainloop()
# if __name__ == "__main__":
#     main()
