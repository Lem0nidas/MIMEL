
from openpyxl import load_workbook
import numpy as np
from copy import copy
import tkinter as tk
from tkinter import filedialog, messagebox, ttk


def check_conditions(func):
    def wrapper(*args, **kwargs):
        if (combobox.get() == "Select an option"):
            messagebox.showwarning(title= "Wrong Input", message="Please choose an option from DropDown list (Monitoring Area)")
            return
        return func(*args, **kwargs)
    return wrapper


@check_conditions
def main():
    wb = load_workbook(origin_entry.get(), data_only=True)
    newDates = int(combobox.get())

    all_targets_dict = dict()

    sheets = wb.sheetnames

    for i in sheets:
        k = 1
        wb.active = wb[i]
        ws = wb.active


        for row_idx in range(ws.max_row, 0, -1):
            row = ws[row_idx]
            if not is_row_empty(row):
                last_non_empty_row = row_idx
                break

        all_dates_dict = dict()

        for k in range(newDates, 0, -1):
            row = ws[last_non_empty_row - k + 1]

            list_for_date = [cell.value for cell in row]
            all_dates_dict[k] = list_for_date


        all_targets_dict[i] = all_dates_dict


    wb.close()


    wb = load_workbook(destination_entry.get())

    sheets = wb.sheetnames

    for i in sheets:
        k = 1
        wb.active = wb[i]
        ws = wb.active

        for row_idx in range(ws.max_row, 0, -1):
            row = ws[row_idx]
            if not is_row_empty(row):
                last_non_empty_row = row_idx
                break

        for k in range(newDates, 0, -1):
            source_row_data = all_targets_dict[i][k]
            max_col = len(source_row_data)

            for col in range(1, max_col + 1):
                cell = ws.cell(row=last_non_empty_row, column=col)
                new_cell = ws.cell(row=last_non_empty_row + 1, column=col)

                new_cell.value = source_row_data[col - 1]

                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

            last_non_empty_row += 1


    save_directory = save_location_entry.get() + "\\NoFormulasExcel.xlsx"
    wb.save(save_directory)

def is_row_empty(row):
    """Check if a row contains only NaN or empty values."""
    for cell in row:
        if cell.value not in (None, '', np.nan):
            return False
    return True

def browse_origin_excel():
    filename = filedialog.askopenfilename(title="Select Raw Data File")
    if filename:
        origin_entry.delete(0, tk.END)
        origin_entry.insert(0, filename)

def browse_destination_excel():
    filename = filedialog.askopenfilename(title="Select Excel File")
    if filename:
        destination_entry.delete(0, tk.END)
        destination_entry.insert(0, filename)

def browse_save_location():
    filename = filedialog.askdirectory(title="Select Save Location")
    if filename:
        save_location_entry.delete(0, tk.END)
        save_location_entry.insert(0, filename)


root = tk.Tk()
root.title("Update Excel file with equations.")
messagebox.showinfo(title="Reminder", message="Before you copy data to excel file, make sure you have a copy of that excel stored somewhere!")


root.geometry("800x300")

root.grid_columnconfigure(0, weight=1)
root.grid_columnconfigure(1, weight=3)
root.grid_columnconfigure(2, weight=1)
root.grid_rowconfigure(0, weight=1)
root.grid_rowconfigure(1, weight=1)
root.grid_rowconfigure(2, weight=1)
root.grid_rowconfigure(3, weight=1)
root.grid_rowconfigure(4, weight=1)
root.grid_rowconfigure(5, weight=1)



origin_label = tk.Label(root, text="Origin Excel")
origin_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")
origin_entry = tk.Entry(root, width=50)
origin_entry.grid(row=0, column=1, padx=10, pady=10, sticky="we")
origin_button = tk.Button(root, text="Browse", command=browse_origin_excel)
origin_button.grid(row=0, column=2, padx=10, pady=10)

destination_label = tk.Label(root, text="Destination Excel")
destination_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")
destination_entry = tk.Entry(root, width=50)
destination_entry.grid(row=1, column=1, padx=10, pady=10, sticky="we")
destination_button = tk.Button(root, text="Browse", command=browse_destination_excel)
destination_button.grid(row=1, column=2, padx=10, pady=10)

save_location_label = tk.Label(root, text="Save Location")
save_location_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")
save_location_entry = tk.Entry(root, width=50)
save_location_entry.grid(row=2, column=1, padx=10, pady=10, sticky="we")
save_location_button = tk.Button(
    root, text="Browse", command=browse_save_location)
save_location_button.grid(row=2, column=2, padx=10, pady=10)

monitoringAreas = ["1", "2", "3", "4", "5", "6", "7", "8"]
dropdown_label = tk.Label(root, text="Number of dates/measurements")
dropdown_label.grid(row=4, column=0, padx=10, pady=10, sticky="w")
combobox = ttk.Combobox(root, values=monitoringAreas, width=20)
combobox.grid(row=4, column=1, padx=10, pady=10, sticky="we")
combobox.set("Select an option")



execute_button = tk.Button(root, text="Execute", command=main)
execute_button.grid(row=5, column=1, padx=10, pady=10, sticky="e")

cancel_button = tk.Button(root, text="Cancel", command=root.quit)
cancel_button.grid(row=5, column=2, padx=10, pady=10, sticky="w")


root.mainloop()